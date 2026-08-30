"""Append the added Digitap-style sheets to the rendered workbook (webapp
download only — the CLI/reference output stays untouched):
Avg Closing Balance (3rd/4th), Spend Analysis, Loan Analysis, Daily Balance."""
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from bankiq.render import _set, _hdr, NAVY, PEACH, BLUE, PINK, AMT, DATE, MONTH, mdt
from openpyxl.styles import Font

from . import insights as I
from .digitap import build_analysis


def _month_headers(ws, months, row, start_col):
    for i, m in enumerate(months):
        _set(ws, f"{get_column_letter(start_col + i)}{row}", mdt(m),
             Font(name="Arial", size=10, bold=True), PINK, MONTH, "center", border=True)


def add_extra_sheets(target, meta, rep):
    """Mutate a live Workbook in place (in-memory flow), or load/save a path."""
    is_wb = isinstance(target, Workbook)
    wb = target if is_wb else load_workbook(target)
    months = rep["months"]

    # 1. Avg Closing Balance 3rd & 4th
    a = I.avg_closing_3_4(rep)
    ws = wb.create_sheet("Avg Closing Bal 3-4")
    _hdr(ws, 1, ["Month", "Closing Balance (3rd)", "Closing Balance (4th)", "Average (3rd & 4th)"],
         [16, 22, 22, 22])
    ws.freeze_panes = "A2"
    r = 2
    for row in a["rows"]:
        _set(ws, f"A{r}", mdt(datetime.date.fromisoformat(row["month"])),
             Font(name="Arial", size=10), fmt=MONTH, border=True)
        _set(ws, f"B{r}", row["close_3"], Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"C{r}", row["close_4"], Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"D{r}", row["avg"], Font(name="Arial", size=10, bold=True), fmt=AMT, border=True)
        r += 1
    _set(ws, f"A{r}", "Overall", Font(name="Arial", size=10, bold=True), PEACH, border=True)
    _set(ws, f"B{r}", a["avg_3"], Font(name="Arial", size=10), BLUE, AMT, border=True)
    _set(ws, f"C{r}", a["avg_4"], Font(name="Arial", size=10), BLUE, AMT, border=True)
    _set(ws, f"D{r}", a["overall_avg"], Font(name="Arial", size=10, bold=True), BLUE, AMT, border=True)

    # 2. Spend Analysis
    rows = I.spend_analysis(meta, months)
    ws = wb.create_sheet("Spend Analysis")
    ncol = len(months)
    _hdr(ws, 1, ["Category"] + [""] * ncol + ["Total", "Count", "% Debits"],
         [26] + [12] * ncol + [16, 9, 10])
    _month_headers(ws, months, 1, 2)
    ws.freeze_panes = "B2"
    r = 2
    for row in rows:
        label = row["category"] + ("  [LIFESTYLE]" if row["lifestyle"] else "")
        _set(ws, f"A{r}", label, Font(name="Arial", size=10, bold=row["lifestyle"]), PEACH, border=True)
        for i, v in enumerate(row["monthly"]):
            _set(ws, f"{get_column_letter(2 + i)}{r}", v or None, Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"{get_column_letter(2 + ncol)}{r}", row["total"], Font(name="Arial", size=10, bold=True), fmt=AMT, border=True)
        _set(ws, f"{get_column_letter(3 + ncol)}{r}", row["count"], Font(name="Arial", size=10), border=True)
        _set(ws, f"{get_column_letter(4 + ncol)}{r}", row["pct_of_debits"] / 100, Font(name="Arial", size=10), fmt="0.0%", border=True)
        r += 1

    # 3. Loan Analysis
    loans = I.loan_analysis(meta)
    ws = wb.create_sheet("Loan Analysis")
    _hdr(ws, 1, ["Lender", "Type", "Pattern", "Txns", "Total Paid", "Monthly Avg", "First Seen", "Last Seen"],
         [30, 14, 22, 8, 16, 14, 14, 14])
    ws.freeze_panes = "A2"
    r = 2
    for l in loans:
        _set(ws, f"A{r}", l["lender"], Font(name="Arial", size=10, bold=True), border=True)
        _set(ws, f"B{r}", l["lender_type"], Font(name="Arial", size=10), border=True)
        _set(ws, f"C{r}", l["pattern"], Font(name="Arial", size=10), border=True)
        _set(ws, f"D{r}", l["txn_count"], Font(name="Arial", size=10), border=True)
        _set(ws, f"E{r}", l["total"], Font(name="Arial", size=10, bold=True), fmt=AMT, border=True)
        _set(ws, f"F{r}", l["monthly_avg"], Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"G{r}", mdt(datetime.date.fromisoformat(l["first_seen"])), Font(name="Arial", size=10), fmt=DATE, border=True)
        _set(ws, f"H{r}", mdt(datetime.date.fromisoformat(l["last_seen"])), Font(name="Arial", size=10), fmt=DATE, border=True)
        r += 1
    _set(ws, f"A{r}", "TOTAL", Font(name="Arial", size=10, bold=True), PEACH, border=True)
    for col in "BC":
        _set(ws, f"{col}{r}", None, border=True)
    _set(ws, f"D{r}", sum(l["txn_count"] for l in loans), Font(name="Arial", size=10, bold=True), BLUE, border=True)
    _set(ws, f"E{r}", round(sum(l["total"] for l in loans), 2), Font(name="Arial", size=10, bold=True), BLUE, AMT, border=True)
    for col in "FGH":
        _set(ws, f"{col}{r}", None, border=True)

    # 4. Daily Balance (Open + Close)
    daily = I.daily_open_close(meta)
    ws = wb.create_sheet("Daily Balance")
    _hdr(ws, 1, ["Date", "Opening Balance", "Closing Balance", "Txns", "Net Change", "Below 1000"],
         [14, 18, 18, 8, 16, 12])
    ws.freeze_panes = "A2"
    for i, d in enumerate(daily):
        r = i + 2
        _set(ws, f"A{r}", mdt(datetime.date.fromisoformat(d["date"])), Font(name="Arial", size=10), fmt=DATE, border=True)
        _set(ws, f"B{r}", d["open"], Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"C{r}", d["close"], Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"D{r}", d["txns"] or None, Font(name="Arial", size=10), border=True)
        _set(ws, f"E{r}", d["net"] or None, Font(name="Arial", size=10), fmt=AMT, border=True)
        _set(ws, f"F{r}", "Y" if d["close"] < 1000 else None, Font(name="Arial", size=10), align="center", border=True)

    # 5. Full Analysis (Digitap parity)
    da = build_analysis(meta, rep)
    ws = wb.create_sheet("Full Analysis")
    ws.freeze_panes = "B2"
    r = 1
    for k, v in da["header"].items():
        _set(ws, f"A{r}", k, Font(name="Arial", size=10, bold=True), PEACH, border=True)
        _set(ws, f"B{r}", v, Font(name="Arial", size=10), border=True)
        r += 1
    r += 1
    ws.column_dimensions["A"].width = 46
    hdr_row = r
    _set(ws, f"A{r}", "Particulars", Font(name="Arial", size=10, bold=True, color="FFFFFFFF"), NAVY, "General", "center", border=True)
    for i, m in enumerate(da["months"]):
        _set(ws, f"{get_column_letter(2 + i)}{r}", mdt(datetime.date.fromisoformat(m)),
             Font(name="Arial", size=10, bold=True), PINK, MONTH, "center", border=True)
        ws.column_dimensions[get_column_letter(2 + i)].width = 15
    ov_col = 2 + len(da["months"])
    _set(ws, f"{get_column_letter(ov_col)}{r}", "Overall", Font(name="Arial", size=10, bold=True), PINK, "General", "center", border=True)
    ws.column_dimensions[get_column_letter(ov_col)].width = 15
    r += 1
    for mrow in da["metrics"]:
        _set(ws, f"A{r}", mrow["label"], Font(name="Arial", size=10), PEACH, border=True)
        for i, v in enumerate(mrow["values"]):
            _set(ws, f"{get_column_letter(2 + i)}{r}", v if v != "" else None,
                 Font(name="Arial", size=10), fmt=(AMT if isinstance(v, float) else "General"), border=True)
        ov = mrow["overall"]
        _set(ws, f"{get_column_letter(ov_col)}{r}", ov if ov != "" else None,
             Font(name="Arial", size=10, bold=True), fmt=(AMT if isinstance(ov, float) else "General"), border=True)
        r += 1

    if is_wb:
        return wb
    wb.save(target)
    return target
