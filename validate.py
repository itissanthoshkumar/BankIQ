#!/usr/bin/env python3
"""Regenerate the three sample statements and diff against the reference
workbooks, sheet by sheet. Reports structural and value-level parity."""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

from bankiq.parse import parse_statement
from bankiq.categorize import categorize_all
from bankiq.analyze import build_report
from bankiq.render import render

DL = "/Users/santhoshp/Downloads"
CASES = [
    (f"{DL}/202885XXXX_DoPW BHUV1205 (2).pdf", "BHUV1205", f"{DL}/BHUVANESHWARI R (1).xlsx"),
    (f"{DL}/Deposit_Statement_535_3438_21_07_2026 (3).pdf", None, f"{DL}/Pandian Selvaraj (1).xlsx"),
    (f"{DL}/SBI Statement Password _ 60382270773 (1).pdf", "60382270773", f"{DL}/VINAYAGAM V.xlsx"),
]

OUTDIR = os.environ.get("BANKIQ_OUT", tempfile.mkdtemp(prefix="bankiq_val_"))


def cells(ws, max_row=None, max_col=None):
    out = {}
    for row in ws.iter_rows(min_row=1, max_row=max_row or ws.max_row,
                            max_col=max_col or ws.max_column):
        for c in row:
            if c.value is not None:
                out[c.coordinate] = c.value
    return out


def norm(v):
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return v


def diff_sheet(wr, wm, name, skip_cells=()):
    """Compare reference sheet vs mine. Returns (n_match, n_diff, samples)."""
    if name not in wm.sheetnames:
        return 0, -1, ["MISSING SHEET"]
    a, b = cells(wr[name]), cells(wm[name])
    keys = set(a) | set(b)
    n_match = n_diff = 0
    samples = []
    import datetime
    for k in sorted(keys, key=lambda s: (len(s), s)):
        if k in skip_cells:
            continue
        va, vb = norm(a.get(k)), norm(b.get(k))
        if isinstance(va, datetime.datetime) and isinstance(vb, datetime.datetime):
            va, vb = va.date(), vb.date()
        if isinstance(va, (int, float)) and isinstance(vb, (int, float)):
            same = abs(va - vb) < 0.005
        else:
            same = va == vb
        if same:
            n_match += 1
        else:
            n_diff += 1
            if len(samples) < 6:
                samples.append(f"{k}: ref={va!r} mine={vb!r}")
    return n_match, n_diff, samples


def main():
    print(f"outputs -> {OUTDIR}")
    for pdf, pw, ref in CASES:
        meta = parse_statement(pdf, pw, tempfile.mkdtemp(prefix="bankiq_"))
        categorize_all(meta)
        rep = build_report(meta)
        out = os.path.join(OUTDIR, os.path.basename(ref).replace(" (1)", "_mine"))
        render(rep, out)
        # formulas compared as strings; literals as values
        wr = load_workbook(ref)
        wm = load_workbook(out)
        print(f"\n================ {meta['bank']} ================")
        print(f"sheets ref={len(wr.sheetnames)} mine={len(wm.sheetnames)}")
        missing = [s for s in wr.sheetnames if s not in wm.sheetnames]
        extra = [s for s in wm.sheetnames if s not in wr.sheetnames]
        if missing or extra:
            print(f"  missing={missing} extra={extra}")
        if wr.sheetnames != wm.sheetnames and not missing and not extra:
            print(f"  ORDER differs: ref={wr.sheetnames}")
        for name in wr.sheetnames:
            skip = set()
            if name == "Analysis":
                skip = {"C7"}          # generated transaction id
            if name == "Statements Considered":
                skip = {"P2"}
            nm, nd, samples = diff_sheet(wr, wm, name, skip)
            total = nm + max(nd, 0)
            pct = 100.0 * nm / total if total else 100.0
            flag = "OK " if nd == 0 else f"{nd:5d} diff"
            print(f"  {name:28s} {flag}  ({pct:5.1f}% of {total})")
            for s in samples:
                print(f"        {s}")


if __name__ == "__main__":
    main()
