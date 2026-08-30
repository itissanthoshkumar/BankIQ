"""Workbook renderer — reproduces the reference report layout exactly."""
import datetime
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

NAVY = "FF003366"
PEACH = "FFFDEADA"
CYAN = "FFE8F5F8"
PINK = "FFF2DCDB"
BLUE = "FFDCE6F2"
YELLOW = "FFFFFF00"
ORANGE = "FFFCD5B5"
LTBLUE = "FFC6D9F1"
SECT = "FFDBEEF4"

AMT = "#,##0.00_);[RED]\\(#,##0.00\\)"
DATE = "d\\-mmm\\-yy"
MONTH = "mmm\\-yy"
SCDATE = "d/mmm/yy"

THIN = Side(style="thin")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

TXN_HEADERS = ["Sl. No. ", "Date", "Cheque No.", "Description", "Amount", "Category", "Balance"]
TXN_WIDTHS = [8.63, 11.88, 9.63, 49.88, 21.88, 30.88, 23.13]


def fill(color):
    return PatternFill("solid", fgColor=color)


def _set(ws, coord, value, font=None, fillc=None, fmt=None, align=None,
         border=False, wrap=False, valign="center"):
    c = ws[coord]
    c.value = value
    c.font = font or Font(name="Arial", size=10)
    if fillc:
        c.fill = fill(fillc)
    if fmt:
        c.number_format = fmt
    c.alignment = Alignment(horizontal=align, vertical=valign, wrap_text=wrap)
    if border:
        c.border = BORDER
    return c


def _hdr(ws, row, headers, widths=None, start_col=1):
    for i, h in enumerate(headers):
        col = start_col + i
        fmt = "#,##0.00" if h in ("Amount", "Balance") else "General"
        _set(ws, f"{get_column_letter(col)}{row}",
             h, Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
             NAVY, fmt, "center", border=True, wrap=True)
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w


def _txn_row(ws, row, t, sl, start_col=1):
    cl = get_column_letter
    vals = [sl, t["date"], t.get("cheque"), t["desc"], t["amount"], t["category"], t["balance"]]
    for i, v in enumerate(vals):
        col = cl(start_col + i)
        c = ws[f"{col}{row}"]
        c.value = datetime.datetime(v.year, v.month, v.day) if isinstance(v, datetime.date) else v
        if i == 1:
            c.number_format = DATE
            c.font = Font(name="Arial", size=10)
        elif i in (4, 6):
            c.number_format = AMT
            c.font = Font(name="Arial", size=10)
        elif i == 5:
            c.font = Font(name="Calibri", size=10, bold=True)
            c.alignment = Alignment(horizontal="left")
            c.number_format = "#,##0.00"
        else:
            c.font = Font(name="Arial", size=10)


def _txn_sheet(wb, name, txns, widths=None):
    ws = wb.create_sheet(name)
    _hdr(ws, 1, TXN_HEADERS, widths or TXN_WIDTHS)
    ws.freeze_panes = "A2"
    for i, t in enumerate(txns):
        _txn_row(ws, i + 2, t, i + 1)
    return ws


def _grouped_sheet(wb, name, groups, widths):
    ws = wb.create_sheet(name)
    _hdr(ws, 1, ["Group"] + TXN_HEADERS, widths)
    ws.freeze_panes = "A2"
    r = 2
    for gno, ts in groups:
        for i, t in enumerate(ts):
            ws[f"A{r}"].value = gno
            ws[f"A{r}"].font = Font(name="Arial", size=10)
            _txn_row(ws, r, t, i + 1, start_col=2)
            r += 1
    return ws


def mdt(d):
    return datetime.datetime(d.year, d.month, d.day)


# ------------------------------------------------------------------ sheets
def sheet_analysis(wb, rep):
    ws = wb.create_sheet("Analysis")
    meta = rep["meta"]
    widths = {"A": 2.63, "B": 49.0}
    for col in "CDEFGHIJKLMNOPQ"[:len(rep["months"]) + 1]:
        widths[col] = 26.5
    for k, v in widths.items():
        ws.column_dimensions[k].width = v

    _set(ws, "A1", "Summary Details", Font(name="Arial", size=14))
    labels = [
        ("Name of the Account Holder", rep["display_name"]),
        ("Address", meta.get("address")),
        ("Email", meta.get("email")),
        ("Mobile Number", meta.get("mobile")),
        ("PAN", meta.get("pan")),
        ("Perfios Transaction Id", rep["perfios_id"]),
        ("Account Number", meta.get("account_no")),
        ("Account Type", meta.get("account_type")),
        ("Derived Account Type", rep["derived_type"]),
    ]
    for i, (lab, val) in enumerate(labels):
        r = 2 + i
        _set(ws, f"B{r}", lab, Font(name="Arial", size=9), PEACH, border=True, wrap=True)
        ws.merge_cells(f"C{r}:E{r}")
        for col in "CDE":
            ws[f"{col}{r}"].border = BORDER
        _set(ws, f"C{r}", val, Font(name="Arial", size=9), CYAN, align="left", border=True)

    _set(ws, "A12", "Monthwise Details", Font(name="Arial", size=14))
    months = rep["months"]
    ncols = len(months)
    total_col = 3 + ncols
    _set(ws, "B13", None, Font(name="Arial", size=10), BLUE, border=True, wrap=True)
    for i, m in enumerate(months):
        _set(ws, f"{get_column_letter(3 + i)}13", mdt(m),
             Font(name="Arial", size=10, bold=True), PINK, MONTH, "center", border=True)
    _set(ws, f"{get_column_letter(total_col)}13", "TOTAL",
         Font(name="Arial", size=10, bold=True), PINK, "General", "center", border=True)
    r = 14
    last_l = get_column_letter(total_col - 1)
    for label, series in rep["monthwise"]:
        fmt = AMT if "(Amount)" in label else "General"
        _set(ws, f"B{r}", label, Font(name="Arial", size=10), PEACH, border=True, wrap=True)
        for i, v in enumerate(series):
            _set(ws, f"{get_column_letter(3 + i)}{r}", v,
                 Font(name="Arial", size=10), BLUE, fmt, "right", border=True)
        _set(ws, f"{get_column_letter(total_col)}{r}", f"=SUM(C{r}:{last_l}{r})",
             Font(name="Arial", size=10), BLUE, "General", "right", border=True)
        r += 1

    _set(ws, "A42", "EOD Balances Table", Font(name="Arial", size=14))
    grid = rep["eod"]
    for j, day in enumerate((7, 14, 21, 28)):
        r = 43 + j
        _set(ws, f"B{r}", f"EOD Balance on {day}th" if day != 21 else "EOD Balance on 21st",
             Font(name="Arial", size=10), PEACH, border=True, wrap=True)
        for i, m in enumerate(months):
            _set(ws, f"{get_column_letter(3 + i)}{r}", grid[m][day - 1],
                 Font(name="Arial", size=10), BLUE, AMT, "right", border=True)
    return ws


def sheet_derived(wb, rep):
    ws = wb.create_sheet("Derived Analysis")
    months = rep["months"]
    ncols = len(months)
    total_col = 3 + ncols
    ws.freeze_panes = "C1"
    ws.column_dimensions["A"].width = 1.13
    ws.column_dimensions["B"].width = 44.75
    for i in range(ncols + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 25.75
    _set(ws, "A1", "Derived Analysis Table", Font(name="Arial", size=14, bold=True), align="left")
    _set(ws, "A3", "Monthwise Details", Font(name="Arial", size=14))
    for i, m in enumerate(months):
        _set(ws, f"{get_column_letter(3 + i)}4", mdt(m),
             Font(name="Arial", size=10, bold=True), PINK, MONTH, "center", border=True)
    _set(ws, f"{get_column_letter(total_col)}4", "TOTAL",
         Font(name="Arial", size=10, bold=True), PINK, "General", "center", border=True)
    last_l = get_column_letter(total_col - 1)
    for j, (label, _series) in enumerate(rep["monthwise"]):
        r = 5 + j
        src = 14 + j
        _set(ws, f"B{r}", label, Font(name="Arial", size=10), PEACH, border=True, wrap=True)
        for i in range(ncols):
            col = get_column_letter(3 + i)
            f = (f"= + IF(OR(('Analysis'!{col}{src}=\"Y\"),('Analysis'!{col}{src}=\"N\")),"
                 f" NA(),'Analysis'!{col}{src})")
            _set(ws, f"{col}{r}", f, Font(name="Arial", size=10), BLUE, AMT, "right", border=True)
        _set(ws, f"{get_column_letter(total_col)}{r}", f"=SUM(C{r}:{last_l}{r})",
             Font(name="Arial", size=10), BLUE, AMT, "right", border=True)
    return ws


def sheet_eod(wb, rep):
    ws = wb.create_sheet("EOD Balance sheet1")
    months = rep["months"]
    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 15.75
    for i in range(len(months)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 25.75
    _set(ws, "A1", "Day/Month", Font(name="Arial", size=10, bold=True), YELLOW,
         "General", "center", border=True)
    for i, m in enumerate(months):
        _set(ws, f"{get_column_letter(2 + i)}1", mdt(m),
             Font(name="Arial", size=10, bold=True), ORANGE, MONTH, "center", border=True)
    for day in range(1, 32):
        r = day + 1
        _set(ws, f"A{r}", day, Font(name="Arial", size=10, bold=True), ORANGE, border=True)
        for i, m in enumerate(months):
            v = rep["eod"][m][day - 1]
            if v is not None:
                _set(ws, f"{get_column_letter(2 + i)}{r}", v, Font(name="Arial", size=10),
                     LTBLUE, AMT, border=True)
    return ws


def sheet_top5(wb, rep, credit=True):
    name = "Top 5 Parties by Credit1" if credit else "Top 5 Parties by Debit1"
    title = "Top N Funds Received" if credit else "Top N Funds Transffered"
    data = rep["top5_credit"] if credit else rep["top5_debit"]
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 1.13
    ws.column_dimensions["B"].width = 20.94 if credit else 19.0
    ws.column_dimensions["C"].width = 24.75
    _set(ws, "A1", title, Font(name="Arial", size=14))
    r = 2
    for m in rep["months"]:
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"C{r}"].border = BORDER
        _set(ws, f"B{r}", mdt(m), Font(name="Arial", size=10, bold=True), PINK,
             MONTH, "center", border=True)
        r += 1
        _set(ws, f"B{r}", "Description", Font(name="Arial", size=10, bold=True),
             YELLOW, "General", "center", border=True)
        _set(ws, f"C{r}", "Amount", Font(name="Arial", size=10, bold=True),
             YELLOW, "General", "center", border=True)
        r += 1
        for desc, amt in data[m]:
            _set(ws, f"B{r}", desc, Font(name="Arial", size=10), PEACH, border=True, wrap=True)
            _set(ws, f"C{r}", round(amt, 2), Font(name="Arial", size=10), BLUE, AMT, border=True)
            r += 1
    return ws


def sheet_statements(wb, rep):
    ws = wb.create_sheet("Statements Considered")
    meta = rep["meta"]
    headers = ["File Name", "Institution", "Account No", "Transaction Start Date",
               "Transaction End Date", "Name as in Statement", "Address as in Statement",
               "Mobile as in Statement", "Landline as in Statement", "Email as in Statement",
               "PAN as in Statement", "Password Protected", "Producer/Creator", "Reason",
               "Statement Status", "Perfios Transaction Id", "Fetched Account Holder Name"]
    _hdr(ws, 1, headers)
    for i in range(len(headers)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 27.25 if i == 15 else 21.25
    ws.freeze_panes = "A2"
    import os
    base = os.path.basename(meta.get("source_file") or "statement.pdf")
    base = re.sub(r"\s*\(\d+\)\s*(?=\.pdf$)", "", base, flags=re.I)
    base = base.replace(" ", "").replace("_", "")
    txns = meta["transactions"]
    start = txns[0]["date"] if txns else None
    end = txns[-1]["date"] if txns else None
    if (meta.get("account_no") or "").upper().startswith("X"):
        fetched = "Masked Account Number"
    else:
        fetched = rep["display_name"]
    vals = [base, meta["institution"], meta.get("account_no"),
            mdt(start) if start else None, mdt(end) if end else None,
            rep["display_name"], meta.get("address"), meta.get("mobile"), None,
            meta.get("email"), meta.get("pan"), meta.get("password_protected"),
            None, None, "VERIFIED", rep["perfios_id"], fetched]
    for i, v in enumerate(vals):
        col = get_column_letter(i + 1)
        if i in (3, 4):
            _set(ws, f"{col}2", v, Font(name="Arial", size=10), fmt=SCDATE, border=True, wrap=True)
        elif i == 14:
            _set(ws, f"{col}2", v, Font(name="Arial", size=11, bold=True, color="0092D050"),
                 align="center", border=True)
        else:
            _set(ws, f"{col}2", v, Font(name="Arial", size=10), border=True, wrap=True)
    return ws


def sheet_fcu(wb, rep):
    ws = wb.create_sheet("FCU Indicators 1")
    trig = rep["fcu"]
    widths = {"A": 6.25, "B": 11.72, "C": 9.75, "D": 43.25, "E": 21.25, "F": 21.68, "G": 22.75}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    cal = lambda bold=False, color="FF002060": Font(name="Calibri", size=11, bold=bold, color=color)

    ws.merge_cells("A1:B1")
    ws["B1"].border = BORDER
    _set(ws, "A1", "FCU Score", Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
         NAVY, "General", "center", border=True)
    _set(ws, "C1", "NA", Font(name="Arial", size=10), align="center", border=True, wrap=True)

    ws.merge_cells("B3:D3")
    for col in "BCD":
        ws[f"{col}3"].border = BORDER
    _set(ws, "B3", "Particulars", Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
         NAVY, "General", "center", border=True, wrap=True)
    _set(ws, "E3", "Possible Fraud", Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
         NAVY, "#,##0.00", "center", border=True, wrap=True)
    _set(ws, "F3", "Behavioural/Transactional", Font(name="Arial", size=10, bold=True, color="FFFFFFFF"),
         NAVY, "#,##0.00", "center", border=True, wrap=True)
    for r, (label, pf, bt) in enumerate(
            [("Number of Triggers Subscribed", 20, 40),
             ("Number of Triggers Identified", 0, len(trig))], start=4):
        _set(ws, f"A{r}", r - 3, Font(name="Arial", size=10), align="center", border=True, wrap=True)
        ws.merge_cells(f"B{r}:D{r}")
        for col in "BCD":
            ws[f"{col}{r}"].border = BORDER
        _set(ws, f"B{r}", label, cal(True, "FF003366"), align="left", border=True, wrap=True)
        _set(ws, f"E{r}", pf, cal(True), fmt="0", align="center", border=True, wrap=True)
        _set(ws, f"F{r}", bt, cal(True), fmt="0", align="center", border=True, wrap=True)

    r = 7
    _set(ws, "A7", "#", cal(True), SECT, align="center", border=True)
    ws.merge_cells("B7:D7")
    for col in "BCD":
        ws[f"{col}7"].border = BORDER
    _set(ws, "B7", "Behavioural/Transactional Indicators", cal(True), SECT,
         align="center", border=True, wrap=True)
    _set(ws, "E7", "Identified?", cal(True), SECT, align="center", border=True, wrap=True)
    _set(ws, "F7", "Remarks", cal(True), SECT, align="center", border=True, wrap=True)

    first_trig_row = 8
    for i, t in enumerate(trig):
        r = first_trig_row + i
        ws.row_dimensions[r].height = 60
        _set(ws, f"A{r}", i + 1, cal(), fmt="0", align="center", border=True)
        ws.merge_cells(f"B{r}:D{r}")
        for col in "BCD":
            ws[f"{col}{r}"].border = BORDER
        _set(ws, f"B{r}", f"{t['title']}\n{t['desc']}", cal(True, "FF003366"),
             align="left", border=True, wrap=True)
        _set(ws, f"E{r}", "Y", cal(), align="center", border=True)
        if isinstance(t["remark"], (int, float)):
            _set(ws, f"F{r}", t["remark"],
                 Font(name="Arial", size=10, color="000000FF"),
                 align="center", border=True, wrap=True)
        else:
            _set(ws, f"F{r}", t["remark"], cal(), align="center", border=True, wrap=True)
    total_r = first_trig_row + len(trig)
    ws.row_dimensions[total_r].height = 15
    _set(ws, f"A{total_r}", None, border=True)
    _set(ws, f"B{total_r}", "Total Count", cal(True), SECT, align="center", border=True, wrap=True)
    _set(ws, f"E{total_r}", f'=COUNTIF(E{first_trig_row}:E{total_r-1},"=Y")',
         cal(), align="center", border=True)
    _set(ws, f"F{total_r}", f"=SUM(F{first_trig_row}:F{total_r-1})",
         cal(), align="center", border=True)

    # detail sections
    r = total_r + 2
    for i, t in enumerate(trig):
        if not t.get("rows") or t.get("section") is None:
            continue
        if t.get("grouped"):
            # irregular EMI lives in its own sheet; hyperlink handled below
            t["_link"] = "'FCU Irregular Xns 1'!A3"
            t["_backrow"] = first_trig_row + i
            continue
        ws.row_dimensions[r].height = 24
        _set(ws, f"A{r}", i + 1, cal(True), SECT, align="center", border=True)
        ws.merge_cells(f"B{r}:D{r}")
        for col in "BCD":
            ws[f"{col}{r}"].border = BORDER
        _set(ws, f"B{r}", t["section"], cal(True), SECT, align="center",
             border=True, wrap=True, valign="bottom")
        r += 1
        _hdr(ws, r, TXN_HEADERS)
        r += 1
        t["_link"] = f"A{r}"
        for j, tx in enumerate(t["rows"]):
            _txn_row(ws, r, tx, j + 1)
            r += 1
        back = ws[f"G{r}"]
        back.value = "BACK"
        back.font = Font(name="Arial", size=10, bold=True)
        back.alignment = Alignment(horizontal="right", vertical="center")
        back.hyperlink = Hyperlink(ref=f"G{r}", location=f"F{first_trig_row + i}")
        r += 1
    # remark hyperlinks
    for i, t in enumerate(trig):
        if isinstance(t["remark"], (int, float)) and t.get("_link"):
            c = ws[f"F{first_trig_row + i}"]
            c.hyperlink = Hyperlink(ref=c.coordinate, location=t["_link"])
    return ws


def sheet_fcu_irregular(wb, rep):
    trig = [t for t in rep["fcu"] if t.get("grouped")]
    if not trig:
        return None
    t = trig[0]
    idx = rep["fcu"].index(t) + 1
    ws = wb.create_sheet("FCU Irregular Xns 1")
    widths = {"A": 6.13, "B": 6.13, "C": 15.25, "D": 9.75, "E": 46.13, "F": 22.88, "G": 14.5, "H": 26.0}
    for k, v in widths.items():
        ws.column_dimensions[k].width = v
    _set(ws, "A1", idx, Font(name="Calibri", size=11, bold=True, color="FF002060"),
         SECT, align="center", border=True)
    ws.merge_cells("B1:E1")
    for col in "BCDE":
        ws[f"{col}1"].border = BORDER
    _set(ws, "B1", "Irregular EMI Transactions",
         Font(name="Calibri", size=11, bold=True, color="FF002060"), SECT,
         align="left", border=True)
    _hdr(ws, 2, ["Group"] + TXN_HEADERS)
    r = 3
    for gno, tx in t["rows"]:
        ws[f"A{r}"].value = gno
        ws[f"A{r}"].font = Font(name="Arial", size=10)
        _txn_row(ws, r, tx, None, start_col=2)
        r += 1
    # per-group serials
    r = 3
    serial = {}
    for gno, _tx in t["rows"]:
        serial[gno] = serial.get(gno, 0) + 1
        ws[f"B{r}"].value = serial[gno]
        r += 1
    back = ws[f"H{r}"]
    back.value = "BACK"
    back.font = Font(name="Arial", size=10, bold=True)
    back.alignment = Alignment(horizontal="right", vertical="center")
    back.hyperlink = Hyperlink(ref=f"H{r}", location=f"'FCU Indicators 1'!F{7 + idx}")
    return ws


def sheet_roundtrip(wb, rep):
    ws = wb.create_sheet("RoundTrippingParties")
    widths = [20.88, 15.63, 24.25, 17.88, 24.63, 26.38]
    headers = ["Party Identified", "# Credits", "Value", "# Debits", "Value", "Difference"]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w
    for i, h in enumerate(headers):
        fmt = "#,##0.00" if h in ("Value", "Difference") else "General"
        _set(ws, f"{get_column_letter(i+1)}1", h,
             Font(name="Arial", size=10, bold=True, color="FFFFFFFF"), NAVY, fmt,
             "center", border=True, wrap=(i == 2 or i == 4 or i == 5))
    for j, (p, nc, vc, nd, vd) in enumerate(rep["round_trip"]):
        r = j + 2
        _set(ws, f"A{r}", p, Font(name="Arial", size=10))
        _set(ws, f"B{r}", nc, Font(name="Arial", size=10))
        _set(ws, f"C{r}", vc, Font(name="Arial", size=10), fmt=AMT)
        _set(ws, f"D{r}", nd, Font(name="Arial", size=10))
        _set(ws, f"E{r}", vd, Font(name="Arial", size=10), fmt=AMT)
        _set(ws, f"F{r}", f"= C{r} - ABS(E{r})", Font(name="Arial", size=10), fmt=AMT)
    return ws


def sheet_party_xns(wb, rep):
    ws = wb.create_sheet("Party Xns")
    meta = rep["meta"]
    headers = ["Sl. No. ", "Bank Name", "Account No.", "Date", "Cheque No.",
               "Description", "Amount", "Category", "Balance"]
    widths = [15.63, 15.63, 22.0, 15.63, 15.63, 46.88, 24.0, 25.25, 22.75]
    for i, h in enumerate(headers):
        fmt = "#,##0.00" if h in ("Amount", "Balance") else "General"
        _set(ws, f"{get_column_letter(i+1)}1", h,
             Font(name="Arial", size=10, bold=True, color="FFFFFFFF"), NAVY, fmt,
             "center", border=True)
        ws.column_dimensions[get_column_letter(i + 1)].width = widths[i]
    r = 2
    for p, txs in rep["party_blocks"]:
        for i, t in enumerate(txs):
            ws[f"A{r}"].value = i + 1
            ws[f"B{r}"].value = meta["institution"]
            ws[f"C{r}"].value = meta.get("account_no")
            c = ws[f"D{r}"]
            c.value = mdt(t["date"])
            c.number_format = DATE
            ws[f"E{r}"].value = t.get("cheque")
            ws[f"F{r}"].value = t["desc"]
            a = ws[f"G{r}"]
            a.value = t["amount"]
            a.number_format = AMT
            cc = ws[f"H{r}"]
            cc.value = t["category"]
            cc.font = Font(name="Calibri", size=10, bold=True)
            cc.alignment = Alignment(horizontal="left")
            b = ws[f"I{r}"]
            b.value = t["balance"]
            b.number_format = AMT
            for col in "ABCEF":
                ws[f"{col}{r}"].font = Font(name="Arial", size=10)
            r += 1
        r += 1  # blank row between parties
    return ws


# ------------------------------------------------------------------ main
def render(rep, out_path=None):
    """Build the workbook. With out_path, save to disk (CLI flow); with
    out_path=None, return the live Workbook for the caller to save in memory."""
    wb = Workbook()
    wb.remove(wb.active)
    sheet_analysis(wb, rep)
    sheet_derived(wb, rep)
    if rep["bounced_rows"]:
        _txn_sheet(wb, "Bounced penal XNS1", rep["bounced_rows"])
    sheet_eod(wb, rep)
    sheet_top5(wb, rep, True)
    sheet_top5(wb, rep, False)
    _txn_sheet(wb, "Transactions1", rep["meta"]["transactions"])
    _txn_sheet(wb, "UPI XNS1", rep["upi"],
               widths=[8.63, 11.88, 9.63, 49.88, 21.88, 29.8, 23.13])
    if rep["loan_disb"]:
        _txn_sheet(wb, "Loan disbursement sheet1", rep["loan_disb"])
    _grouped_sheet(wb, "Recurring credits1", rep["recurring_credits"],
                   [6.13, 6.13, 15.25, 9.63, 46.13, 22.88, 25.25, 26.0])
    _grouped_sheet(wb, "Recurring debits1", rep["recurring_debits"],
                   [6.13, 6.13, 15.25, 9.63, 46.13, 22.88, 23.5, 26.0])
    _txn_sheet(wb, "High value Credit1", rep["hv_credit"],
               widths=[8.63, 11.88, 9.63, 49.88, 21.88, 25.25, 23.13])
    _txn_sheet(wb, "High value debit1", rep["hv_debit"],
               widths=[8.63, 11.88, 9.63, 49.88, 21.88, 22.77, 23.13])
    if rep["salary_rows"]:
        _txn_sheet(wb, "Salary Xns Sheet1", rep["salary_rows"])
    sheet_statements(wb, rep)
    sheet_fcu(wb, rep)
    sheet_fcu_irregular(wb, rep)
    sheet_roundtrip(wb, rep)
    sheet_party_xns(wb, rep)
    if out_path is None:
        return wb
    wb.save(out_path)
    return out_path
