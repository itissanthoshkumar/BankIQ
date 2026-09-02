#!/usr/bin/env python3
"""Digitap parity: compare our Full-Analysis metric grid against Digitap's own
Analysis sheet, metric by metric and month by month.

Run:  python3 tests/parity_digitap.py
"""
import os
import re
import sys

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from bankiq.parse import parse_statement           # noqa: E402
from bankiq.categorize import categorize_all        # noqa: E402
from bankiq.analyze import build_report             # noqa: E402
from webapp.digitap import build_analysis           # noqa: E402

D = os.path.expanduser("~/Downloads")
REF = f"{D}/report_test123_da4f6c72e"
CASES = [
    ("Union / BHUVANESHWARI", f"{D}/202885XXXX_DoPW BHUV1205 (2).pdf", "BHUV1205",
     f"{REF}/report_test123_da4f6be22.xlsx"),
    ("CUB / PANDIAN", f"{D}/Deposit_Statement_535_3438_21_07_2026 (3).pdf", None,
     f"{REF}/report_test123_da4f6c72e.xlsx"),
    ("SBI / VINAYAGAM", f"{D}/SBI Statement Password _ 60382270773 (1).pdf", "60382270773",
     f"{REF}/report_test123_da4f6c7e7.xlsx"),
]
MONTHS = ("january february march april may june july august september "
          "october november december").split()


def num(v):
    if v is None or v == "":
        return None
    try:
        return round(float(str(v).replace(",", "")), 2)
    except ValueError:
        return None


def ref_grid(path):
    """{metric_label_lower: {month_key: value}} from Digitap's Analysis sheet."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Analysis"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    hdr_i = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Particulars")
    cols = {}
    for j, cell in enumerate(rows[hdr_i][1:], start=1):
        s = str(cell or "").strip().lower()
        m = re.match(r"([a-z]+)\s+(\d{4})", s)
        if m and m.group(1) in MONTHS:
            cols[j] = f"{m.group(2)}-{MONTHS.index(m.group(1)) + 1:02d}"
        elif s == "overall":
            cols[j] = "overall"
    out = {}
    for r in rows[hdr_i + 1:]:
        if not r or not r[0]:
            continue
        out[str(r[0]).strip().lower()] = {k: num(r[j]) for j, k in cols.items() if j < len(r)}
    return out


def main():
    grand_ok = grand_cmp = 0
    for label, pdf, pw, ref in CASES:
        if not (os.path.exists(pdf) and os.path.exists(ref)):
            print(f"{label}: reference or statement missing — skipped")
            continue
        m = parse_statement(pdf, password=pw)
        categorize_all(m)
        rep = build_report(m)
        d = build_analysis(m, rep)
        mine = {}
        for met in d["metrics"]:
            col = {f"{mo[:7]}": num(v) for mo, v in zip(d["months"], met["values"])}
            col["overall"] = num(met["overall"])
            mine[met["label"].strip().lower()] = col
        R = ref_grid(ref)

        shared = [k for k in R if k in mine]
        ok = cmp = 0
        misses = []
        for k in shared:
            for mo, rv in R[k].items():
                mv = mine[k].get(mo)
                if rv is None or mv is None:
                    continue          # blank on either side isn't a disagreement
                cmp += 1
                if abs(rv - mv) <= max(0.02, abs(rv) * 0.005):
                    ok += 1
                elif len(misses) < 6:
                    misses.append(f"{k[:44]} [{mo}] digitap={rv} mine={mv}")
        grand_ok += ok
        grand_cmp += cmp
        pct = 100 * ok / cmp if cmp else 0
        print(f"{label:24} metrics matched {len(shared):>3}/{len(R)}  "
              f"values {ok}/{cmp} = {pct:.1f}%")
        for x in misses:
            print(f"      · {x}")
    if grand_cmp:
        print(f"\nDIGITAP PARITY OVERALL: {grand_ok}/{grand_cmp} = {100*grand_ok/grand_cmp:.1f}%")
    return 0


if __name__ == "__main__":
    sys.exit(main())
